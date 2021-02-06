
# wb=xl.load_workbook("hello.xlsx")
import os
from cert import PrintCert
path = 'stud.xlsx'
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


wb=load_workbook(filename=path)
ws = wb.active

print(ws)
# for cell in wb['b2':'b10']:
# 	# for cell in row:
# 	# 	print(cell.value)
# 	print(cell.value)
# for row in openpyxl.row:
# 	print(row)
# PrintCert("Kanika Naik")
# m_row = len(ws['b'])
m_row=1
for i in range(2,m_row+1):
	row1=ws.cell(row=i,column=2)
	names = row1.value
	PrintCert(names)
wb.close()




